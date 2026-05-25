"""Excel 导出 — 4-sheet 评测报告生成器。

迁移自 W1 时期 `multi_turn_eval/report.py::build_excel_report`，
改为直接从平台 DB（SQLAlchemy session）拉数据，而非读 JSON 文件。

Sheet 结构：
  1. 总览 — 元信息、加权总分、维度权重表
  2. 维度汇总 — 各维度均值/min/max/通过率/样本数
  3. 会话明细 — 每行一个 case，含 6 维分数、最低维度
  4. 轮次明细 — 从 eval_turn_result 或 dim_results_full 反推每轮分数
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.config import DEFAULT_DIMENSION_WEIGHTS, DIMENSION_NAMES
from app.models import (
    BotRewrite,
    BotVersion,
    Conversation,
    Dataset,
    EvalCaseResult,
    EvalRun,
    EvalTurnResult,
    JudgeModel,
    Turn,
)
from app.services.scoring import aggregate_dimension_summary

DIM_CODES = ["dim1", "dim2", "dim3", "dim4", "dim5", "dim6"]
TURN_LEVEL_DIMS = ("dim1", "dim3", "dim4", "dim5")
SESSION_LEVEL_DIMS = ("dim2", "dim6")

# Per-Turn / Session sheet cell 中 reasoning 字段的最长字符数（避免单格超长破坏 xlsx）
_MAX_REASONING_LEN = 500


def _truncate(s: Any, limit: int = _MAX_REASONING_LEN) -> str:
    if s is None or s == "" or s == [] or s == {}:
        return ""
    if isinstance(s, (list, dict)):
        try:
            import json as _json
            t = _json.dumps(s, ensure_ascii=False)
        except Exception:  # noqa: BLE001
            t = str(s)
    else:
        t = str(s)
    return t if len(t) <= limit else t[:limit] + "…"


def _extract_reasoning(judge_raw: Any) -> str:
    """从 judge_raw_response（或 dim_results_full[dim]）抽 reasoning 文本。
    优先级 explanation > issue > note > error，全部空时回退到 JSON 片段。
    """
    if not isinstance(judge_raw, dict):
        return _truncate(judge_raw) if judge_raw else ""
    for key in ("explanation", "issue", "note", "error"):
        v = judge_raw.get(key)
        if v:
            return _truncate(v)
    # 兜底：把整个 dict json dump 截短
    try:
        import json as _json
        return _truncate(_json.dumps(judge_raw, ensure_ascii=False))
    except Exception:  # noqa: BLE001
        return _truncate(str(judge_raw))


# ---------- 样式工具 ----------

_title_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
_header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
_body_font = Font(name="微软雅黑", size=10)
_title_fill = PatternFill("solid", fgColor="305496")
_header_fill = PatternFill("solid", fgColor="4472C4")
_low_fill = PatternFill("solid", fgColor="F8CBAD")
_mid_fill = PatternFill("solid", fgColor="FFF2CC")
_good_fill = PatternFill("solid", fgColor="C6E0B4")
_thin_border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
_left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, row_idx: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _center
        cell.border = _thin_border


def _score_fill(score: float | None):
    if score is None:
        return None
    if score < 0.6:
        return _low_fill
    if score < 0.8:
        return _mid_fill
    return _good_fill


def _fmt(v: Any) -> Any:
    """None → '-'，float 保留 4 位。"""
    if v is None:
        return "-"
    if isinstance(v, float):
        return round(v, 4)
    return v


# ---------- 主入口 ----------

def export_eval_run_xlsx(db: Session, eval_run_id: int) -> bytes:
    """生成 4-sheet Excel report 字节流。

    Raises:
        ValueError: run 不存在
    """
    run = db.get(EvalRun, eval_run_id)
    if not run:
        raise ValueError(f"eval run {eval_run_id} not found")

    cases = (
        db.query(EvalCaseResult)
        .filter(EvalCaseResult.eval_run_id == eval_run_id)
        .all()
    )

    # 预拉关联实体
    dataset = db.get(Dataset, run.dataset_id) if run.dataset_id else None
    bot = db.get(BotVersion, run.bot_version_id) if run.bot_version_id else None
    judge = db.get(JudgeModel, run.judge_model_id) if run.judge_model_id else None

    # conversation_id → conversation_id_src 映射（用于会话明细可读化）
    conv_ids = [c.conversation_id for c in cases]
    conv_src_map: dict[int, str] = {}
    if conv_ids:
        for conv in db.query(Conversation).filter(Conversation.id.in_(conv_ids)).all():
            conv_src_map[conv.id] = conv.conversation_id_src

    dims_selected = run.dimensions_selected or DIM_CODES
    # 仅用 DIM_CODES 的全集做表头，但根据 dims_selected 决定哪些列有意义
    case_dicts = [
        {f"{code}_score": getattr(c, f"{code}_score") for code in DIM_CODES}
        for c in cases
    ]
    dim_summary = aggregate_dimension_summary(case_dicts, DIM_CODES)
    summary_by_code = {row["dimension_code"]: row for row in dim_summary}

    wb = Workbook()

    _build_sheet_overview(wb, run, dataset, bot, judge, len(cases), summary_by_code)
    _build_sheet_dim_summary(wb, summary_by_code)
    # A.4：新 Per-Turn / Session 两 sheet（按 query/会话粒度宽表）
    _build_sheet_per_turn_detail(db, wb, run, cases, conv_src_map)
    _build_sheet_session_detail(wb, cases, conv_src_map)
    # 保留旧"会话明细"sheet 便于看总分一栏对照
    _build_sheet_case_detail(wb, cases, conv_src_map)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- Sheet 1: 总览 ----------

def _build_sheet_overview(wb, run, dataset, bot, judge, case_count, summary_by_code) -> None:
    ws = wb.active
    ws.title = "总览"

    ws.merge_cells("A1:D1")
    ws["A1"] = "多轮对话评测报告 — 总览"
    ws["A1"].font = _title_font
    ws["A1"].fill = _title_fill
    ws["A1"].alignment = _center

    info_rows = [
        ("评测名称", run.name),
        ("描述", run.description or "-"),
        ("状态", run.status),
        ("创建时间", run.created_at.strftime("%Y-%m-%d %H:%M:%S") if run.created_at else "-"),
        ("开始时间", run.started_at.strftime("%Y-%m-%d %H:%M:%S") if run.started_at else "-"),
        ("结束时间", run.finished_at.strftime("%Y-%m-%d %H:%M:%S") if run.finished_at else "-"),
        ("Dataset", f"#{dataset.id} · {dataset.name}" if dataset else f"#{run.dataset_id}"),
        ("Bot 版本", f"#{bot.id} · {bot.name} ({bot.version_tag})" if bot else f"#{run.bot_version_id}"),
        ("Judge 模型", f"#{judge.id} · {judge.name} ({judge.provider}/{judge.model_id})" if judge else f"#{run.judge_model_id}"),
        ("采样数 / 总数", f"{run.sampling_count or '全量'} / {run.total}"),
        ("已完成", run.completed),
        ("失败", run.failed),
        ("加权总分", _fmt(run.weighted_score)),
        ("通过率", f"{run.pass_rate * 100:.1f}%" if run.pass_rate is not None else "-"),
    ]
    r = 3
    for label, val in info_rows:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(bold=True)
        c1.alignment = _left
        ws.cell(row=r, column=2, value=val).alignment = _left
        r += 1

    # 维度权重表
    r += 1
    headers = ["维度", "代码", "权重", "平均分", "加权贡献"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, len(headers))
    r += 1
    for code in DIM_CODES:
        name = DIMENSION_NAMES.get(code, code)
        weight = DEFAULT_DIMENSION_WEIGHTS.get(code, 0.0)
        avg = summary_by_code.get(code, {}).get("avg_score")
        contrib = round(avg * weight, 4) if avg is not None else None
        ws.cell(row=r, column=1, value=name).alignment = _left
        ws.cell(row=r, column=2, value=code).alignment = _center
        ws.cell(row=r, column=3, value=f"{int(weight * 100)}%").alignment = _center
        c_avg = ws.cell(row=r, column=4, value=_fmt(avg))
        c_avg.alignment = _center
        fill = _score_fill(avg)
        if fill:
            c_avg.fill = fill
            if avg is not None and avg < 0.6:
                c_avg.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)
        ws.cell(row=r, column=5, value=_fmt(contrib)).alignment = _center
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = _thin_border
            if ws.cell(row=r, column=c).font is None or ws.cell(row=r, column=c).font.color is None:
                ws.cell(row=r, column=c).font = _body_font
        r += 1

    for col, w in zip("ABCDE", [22, 14, 12, 14, 14]):
        ws.column_dimensions[col].width = w


# ---------- Sheet 2: 维度汇总 ----------

def _build_sheet_dim_summary(wb, summary_by_code) -> None:
    ws = wb.create_sheet("维度汇总")
    headers = ["维度", "代码", "权重", "样本数", "平均分", "最低分", "最高分", "通过率(>=0.6)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    r = 2
    for code in DIM_CODES:
        name = DIMENSION_NAMES.get(code, code)
        weight = DEFAULT_DIMENSION_WEIGHTS.get(code, 0.0)
        row = summary_by_code.get(code, {})
        avg = row.get("avg_score")
        pass_r = row.get("pass_rate")

        ws.cell(row=r, column=1, value=name).alignment = _left
        ws.cell(row=r, column=2, value=code).alignment = _center
        ws.cell(row=r, column=3, value=f"{int(weight * 100)}%").alignment = _center
        ws.cell(row=r, column=4, value=row.get("sample_count", 0)).alignment = _center
        c_avg = ws.cell(row=r, column=5, value=_fmt(avg))
        c_avg.alignment = _center
        fill = _score_fill(avg)
        if fill:
            c_avg.fill = fill
        if avg is not None and avg < 0.6:
            c_avg.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)
        ws.cell(row=r, column=6, value=_fmt(row.get("min_score"))).alignment = _center
        ws.cell(row=r, column=7, value=_fmt(row.get("max_score"))).alignment = _center
        ws.cell(
            row=r,
            column=8,
            value=f"{pass_r * 100:.1f}%" if pass_r is not None else "-",
        ).alignment = _center

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = _thin_border
            if c != 5 or avg is None or avg >= 0.6:
                ws.cell(row=r, column=c).font = _body_font
        r += 1

    for i, w in enumerate([20, 10, 10, 10, 12, 12, 12, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- Sheet 3: 会话明细 ----------

def _build_sheet_case_detail(wb, cases, conv_src_map) -> None:
    ws = wb.create_sheet("会话明细")
    headers = (
        ["case_id", "会话ID", "源会话ID", "加权得分"]
        + [DIMENSION_NAMES.get(code, code) for code in DIM_CODES]
        + ["最低维度", "最低分", "错误"]
    )
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    r = 2
    # 按加权得分升序，低分在前
    sorted_cases = sorted(
        cases,
        key=lambda c: (c.weighted_score is None, c.weighted_score if c.weighted_score is not None else 0),
    )
    for case in sorted_cases:
        ws.cell(row=r, column=1, value=case.id).alignment = _center
        ws.cell(row=r, column=2, value=case.conversation_id).alignment = _center
        ws.cell(row=r, column=3, value=conv_src_map.get(case.conversation_id, "-")).alignment = _center

        c_w = ws.cell(row=r, column=4, value=_fmt(case.weighted_score))
        c_w.alignment = _center
        fill = _score_fill(case.weighted_score)
        if fill:
            c_w.fill = fill

        dim_scores = {code: getattr(case, f"{code}_score") for code in DIM_CODES}
        for j, code in enumerate(DIM_CODES):
            v = dim_scores[code]
            cell = ws.cell(row=r, column=5 + j, value=_fmt(v))
            cell.alignment = _center
            cf = _score_fill(v)
            if cf:
                cell.fill = cf
            if v is not None and v < 0.6:
                cell.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)

        valid_dims = {k: v for k, v in dim_scores.items() if v is not None}
        if valid_dims:
            min_code = min(valid_dims, key=lambda k: valid_dims[k])
            ws.cell(row=r, column=5 + len(DIM_CODES), value=DIMENSION_NAMES.get(min_code, min_code)).alignment = _center
            ws.cell(row=r, column=6 + len(DIM_CODES), value=round(valid_dims[min_code], 4)).alignment = _center
        else:
            ws.cell(row=r, column=5 + len(DIM_CODES), value="-").alignment = _center
            ws.cell(row=r, column=6 + len(DIM_CODES), value="-").alignment = _center

        ws.cell(row=r, column=7 + len(DIM_CODES), value=(case.error or "-")[:200]).alignment = _left

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = _thin_border
            if ws.cell(row=r, column=c).font is None or not getattr(ws.cell(row=r, column=c).font, "color", None):
                ws.cell(row=r, column=c).font = _body_font
        r += 1

    ws.freeze_panes = "E2"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    for i in range(5, 5 + len(DIM_CODES)):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions[get_column_letter(5 + len(DIM_CODES))].width = 18
    ws.column_dimensions[get_column_letter(6 + len(DIM_CODES))].width = 12
    ws.column_dimensions[get_column_letter(7 + len(DIM_CODES))].width = 40


# ---------- Sheet 3: Per-Turn 维度明细 (A.4 — 每 query 一行的宽表) ----------


def _gather_turn_metadata(
    db: Session, run: EvalRun, cases: list[EvalCaseResult]
) -> dict[tuple[int, int], dict]:
    """一次性拉所有 (case_id, turn_index) → {user_query, rewritten_query, bot_response, intent_type, timestamp}。

    走 EvalCaseResult.conversation_id → Conversation.turns → BotRewrite。
    """
    if not cases:
        return {}
    conv_ids = list({c.conversation_id for c in cases})
    bot_version_id = run.bot_version_id
    rows = (
        db.query(
            Turn.conversation_id,
            Turn.turn_index,
            Turn.user_query,
            Turn.timestamp,
            BotRewrite.rewritten_query,
            BotRewrite.bot_response,
            BotRewrite.intent_type,
        )
        .outerjoin(
            BotRewrite,
            (BotRewrite.turn_id == Turn.id)
            & (BotRewrite.bot_version_id == bot_version_id),
        )
        .filter(Turn.conversation_id.in_(conv_ids))
        .order_by(Turn.conversation_id, Turn.turn_index)
        .all()
    )
    # 按 conversation_id 聚合
    by_conv: dict[int, dict[int, dict]] = {}
    for cid, tidx, uq, ts, rq, br, intent in rows:
        by_conv.setdefault(cid, {})[tidx] = {
            "user_query": uq or "",
            "rewritten_query": rq or "",
            "bot_response": br or "",
            "intent_type": intent or "",
            "timestamp": ts or "",
        }
    # 转换为 (case_id, turn_index) 索引
    out: dict[tuple[int, int], dict] = {}
    for case in cases:
        per = by_conv.get(case.conversation_id, {})
        for tidx, meta in per.items():
            out[(case.id, tidx)] = meta
    return out


def _gather_turn_results(db: Session, cases: list[EvalCaseResult]) -> dict[tuple[int, int, str], EvalTurnResult]:
    """(case_id, turn_index, dim_code) → EvalTurnResult。"""
    if not cases:
        return {}
    case_ids = [c.id for c in cases]
    rows = (
        db.query(EvalTurnResult)
        .filter(EvalTurnResult.eval_case_result_id.in_(case_ids))
        .all()
    )
    out: dict[tuple[int, int, str], EvalTurnResult] = {}
    for tr in rows:
        out[(tr.eval_case_result_id, tr.turn_index, tr.dimension_code)] = tr
    return out


def _build_sheet_per_turn_detail(
    db: Session, wb, run: EvalRun, cases: list[EvalCaseResult], conv_src_map: dict[int, str]
) -> None:
    """每 (case, turn) 一行的宽表：含 user/rewrite/bot_response/intent + 4 个 turn-level 维度的分与 reasoning。

    维度顺序：dim1（无 applicable）/ dim3 / dim4 / dim5。
    """
    ws = wb.create_sheet("Per-Turn 维度明细")
    headers: list[str] = [
        "case_id", "源会话ID", "轮次", "时间戳",
        "用户query (原始)", "改写query", "bot回复", "bot意图",
        # dim1：无 applicable
        "改写忠实性 得分", "改写忠实性 reasoning",
        # dim3
        "意图边界识别 适用", "意图边界识别 类型", "意图边界识别 得分", "意图边界识别 reasoning",
        # dim4
        "指代消解 适用", "指代消解 类型", "指代消解 得分", "指代消解 reasoning",
        # dim5
        "重复请求 适用", "重复请求 得分", "重复请求 reasoning",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    if not cases:
        ws.freeze_panes = "E2"
        return

    turn_meta = _gather_turn_metadata(db, run, cases)
    turn_results = _gather_turn_results(db, cases)

    # 用 (case_id, turn_index) 排序写入
    keys = sorted(turn_meta.keys())
    r = 2
    for case_id, turn_index in keys:
        meta = turn_meta[(case_id, turn_index)]
        # 无 user_query 的 phantom-only turn 也展示一行（reasoning 列为 "-"）
        case = next((c for c in cases if c.id == case_id), None)
        src_id = conv_src_map.get(case.conversation_id, "-") if case else "-"

        col = 1
        ws.cell(row=r, column=col, value=case_id).alignment = _center; col += 1
        ws.cell(row=r, column=col, value=src_id).alignment = _center; col += 1
        ws.cell(row=r, column=col, value=turn_index).alignment = _center; col += 1
        ws.cell(row=r, column=col, value=meta["timestamp"]).alignment = _center; col += 1
        ws.cell(row=r, column=col, value=_truncate(meta["user_query"], 300)).alignment = _left; col += 1
        ws.cell(row=r, column=col, value=_truncate(meta["rewritten_query"], 300)).alignment = _left; col += 1
        ws.cell(row=r, column=col, value=_truncate(meta["bot_response"], 300)).alignment = _left; col += 1
        ws.cell(row=r, column=col, value=meta["intent_type"] or "-").alignment = _center; col += 1

        for dim_code in TURN_LEVEL_DIMS:
            tr = turn_results.get((case_id, turn_index, dim_code))
            score = tr.score if tr else None
            applicable = tr.applicable if tr else None
            reasoning = _extract_reasoning(tr.judge_raw_response if tr else None)
            raw = tr.judge_raw_response if tr and isinstance(tr.judge_raw_response, dict) else {}

            if dim_code != "dim1":
                # applicable
                ws.cell(
                    row=r, column=col,
                    value=("是" if applicable else "否") if applicable is not None else "-",
                ).alignment = _center
                col += 1
            if dim_code == "dim3":
                ws.cell(row=r, column=col, value=raw.get("boundary_type") or "-").alignment = _center
                col += 1
            elif dim_code == "dim4":
                ws.cell(row=r, column=col, value=raw.get("anaphora_type") or "-").alignment = _center
                col += 1
            # 得分
            sc = ws.cell(row=r, column=col, value=_fmt(score))
            sc.alignment = _center
            cf = _score_fill(score)
            if cf:
                sc.fill = cf
            if score is not None and score < 0.6:
                sc.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)
            col += 1
            # reasoning
            ws.cell(row=r, column=col, value=reasoning).alignment = _left
            col += 1

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = _thin_border
            if ws.cell(row=r, column=c).font is None or not getattr(
                ws.cell(row=r, column=c).font, "color", None
            ):
                ws.cell(row=r, column=c).font = _body_font
        r += 1

    ws.freeze_panes = "E2"
    # 列宽（保持可读）
    width_map = {
        "A": 8, "B": 18, "C": 6, "D": 18,
        "E": 35, "F": 35, "G": 35, "H": 12,
        "I": 12, "J": 40,                     # dim1
        "K": 8, "L": 14, "M": 10, "N": 36,    # dim3
        "O": 8, "P": 14, "Q": 10, "R": 36,    # dim4
        "S": 8, "T": 10, "U": 36,             # dim5
    }
    for col, w in width_map.items():
        ws.column_dimensions[col].width = w


# ---------- Sheet 4: Session 维度明细 (A.4 — 每会话一行，dim2/6 + turn 级均值) ----------


def _build_sheet_session_detail(wb, cases: list[EvalCaseResult], conv_src_map: dict[int, str]) -> None:
    """每个会话一行：含 dim2/dim6 session-level 得分 + reasoning + 关键证据；
    并附 dim1/3/4/5 的均值（便于一眼看会话级表现）。
    """
    ws = wb.create_sheet("Session 维度明细")
    headers = [
        "case_id", "源会话ID", "轮次数", "加权得分", "最低维度",
        # dim2
        "跨轮记忆保留 得分", "跨轮记忆保留 reasoning",
        "dim2 bot伪声明 false_inherited",
        "dim2 漏报约束 missed_constraints",
        "dim2 precision", "dim2 recall",
        # dim6
        "纠错响应 适用", "纠错响应 得分", "纠错响应 reasoning",
        "dim6 correction_signals",
        # 其它维度均值
        "改写忠实性 均值", "意图边界 均值", "指代消解 均值", "重复请求 均值",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    r = 2
    sorted_cases = sorted(
        cases,
        key=lambda c: (c.weighted_score is None, c.weighted_score if c.weighted_score is not None else 0),
    )
    for case in sorted_cases:
        full = case.dim_results_full or {}
        if not isinstance(full, dict):
            full = {}
        src_id = conv_src_map.get(case.conversation_id, "-")

        dim2 = full.get("dim2") if isinstance(full.get("dim2"), dict) else {}
        dim6 = full.get("dim6") if isinstance(full.get("dim6"), dict) else {}
        # dim2/6 的真实 detail 在 dim_results_full[dim_code]["detail"] 里
        dim2_detail = dim2.get("detail") if isinstance(dim2.get("detail"), dict) else dim2
        dim6_detail = dim6.get("detail") if isinstance(dim6.get("detail"), dict) else dim6

        col = 1
        ws.cell(row=r, column=col, value=case.id).alignment = _center; col += 1
        ws.cell(row=r, column=col, value=src_id).alignment = _center; col += 1
        # 轮次数：从 dim_results_full 推不出，从 case.dim_results_full 一般也无
        # 这里走从 case.turn_results 估（即 dim1 个数）
        ws.cell(row=r, column=col, value="-").alignment = _center; col += 1
        c_w = ws.cell(row=r, column=col, value=_fmt(case.weighted_score))
        c_w.alignment = _center
        fill = _score_fill(case.weighted_score)
        if fill:
            c_w.fill = fill
        col += 1
        ws.cell(
            row=r, column=col,
            value=DIMENSION_NAMES.get(case.lowest_dim_code, case.lowest_dim_code) if case.lowest_dim_code else "-",
        ).alignment = _center
        col += 1

        # dim2 columns
        c_d2 = ws.cell(row=r, column=col, value=_fmt(case.dim2_score))
        c_d2.alignment = _center
        cf = _score_fill(case.dim2_score)
        if cf:
            c_d2.fill = cf
        if case.dim2_score is not None and case.dim2_score < 0.6:
            c_d2.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)
        col += 1
        ws.cell(row=r, column=col, value=_extract_reasoning(dim2_detail)).alignment = _left
        col += 1
        ws.cell(
            row=r, column=col,
            value=_truncate(dim2_detail.get("false_inherited") if isinstance(dim2_detail, dict) else "", 400),
        ).alignment = _left
        col += 1
        ws.cell(
            row=r, column=col,
            value=_truncate(dim2_detail.get("missed_constraints") if isinstance(dim2_detail, dict) else "", 400),
        ).alignment = _left
        col += 1
        ws.cell(
            row=r, column=col,
            value=_fmt(dim2_detail.get("precision") if isinstance(dim2_detail, dict) else None),
        ).alignment = _center
        col += 1
        ws.cell(
            row=r, column=col,
            value=_fmt(dim2_detail.get("recall") if isinstance(dim2_detail, dict) else None),
        ).alignment = _center
        col += 1

        # dim6 columns
        dim6_appl = dim6_detail.get("applicable") if isinstance(dim6_detail, dict) else None
        ws.cell(
            row=r, column=col,
            value=("是" if dim6_appl else "否") if dim6_appl is not None else "-",
        ).alignment = _center
        col += 1
        c_d6 = ws.cell(row=r, column=col, value=_fmt(case.dim6_score))
        c_d6.alignment = _center
        cf = _score_fill(case.dim6_score)
        if cf:
            c_d6.fill = cf
        col += 1
        ws.cell(row=r, column=col, value=_extract_reasoning(dim6_detail)).alignment = _left
        col += 1
        ws.cell(
            row=r, column=col,
            value=_truncate(dim6_detail.get("correction_signals") if isinstance(dim6_detail, dict) else "", 300),
        ).alignment = _left
        col += 1

        # 其它维度均值（直接取已 aggregated 的 case 字段）
        for code in ("dim1", "dim3", "dim4", "dim5"):
            v = getattr(case, f"{code}_score")
            cell = ws.cell(row=r, column=col, value=_fmt(v))
            cell.alignment = _center
            cf = _score_fill(v)
            if cf:
                cell.fill = cf
            if v is not None and v < 0.6:
                cell.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)
            col += 1

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = _thin_border
            if ws.cell(row=r, column=c).font is None or not getattr(
                ws.cell(row=r, column=c).font, "color", None
            ):
                ws.cell(row=r, column=c).font = _body_font
        r += 1

    ws.freeze_panes = "C2"
    widths = [8, 18, 8, 12, 16,
              12, 40, 35, 35, 10, 10,
              10, 12, 36, 30,
              12, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# Markdown export (C.1)
# ============================================================================


def _md_pct(v: float | None) -> str:
    return f"{v * 100:.1f}%" if v is not None else "-"


def _md_num(v: Any, digits: int = 4) -> str:
    if v is None:
        return "-"
    if isinstance(v, float):
        return f"{round(v, digits)}"
    return str(v)


def _md_escape(s: str | None) -> str:
    """转义 Markdown 表格中破坏排版的字符（| 与换行）。"""
    if s is None:
        return "-"
    return str(s).replace("|", "\\|").replace("\n", " ").replace("\r", " ")


def export_eval_run_md(db: Session, eval_run_id: int) -> bytes:
    """Markdown 4-section report，UTF-8 编码字节流。

    Section: 标题/元信息 → 维度汇总表 → 会话明细表（低分前若干）→ 低分前 5 case 备注
    """
    run = db.get(EvalRun, eval_run_id)
    if not run:
        raise ValueError(f"eval run {eval_run_id} not found")

    cases = (
        db.query(EvalCaseResult)
        .filter(EvalCaseResult.eval_run_id == eval_run_id)
        .all()
    )
    dataset = db.get(Dataset, run.dataset_id) if run.dataset_id else None
    bot = db.get(BotVersion, run.bot_version_id) if run.bot_version_id else None
    judge = db.get(JudgeModel, run.judge_model_id) if run.judge_model_id else None

    conv_ids = [c.conversation_id for c in cases]
    conv_src_map: dict[int, str] = {}
    if conv_ids:
        for conv in db.query(Conversation).filter(Conversation.id.in_(conv_ids)).all():
            conv_src_map[conv.id] = conv.conversation_id_src

    case_dicts = [
        {f"{code}_score": getattr(c, f"{code}_score") for code in DIM_CODES}
        for c in cases
    ]
    dim_summary = aggregate_dimension_summary(case_dicts, DIM_CODES)
    summary_by_code = {row["dimension_code"]: row for row in dim_summary}

    lines: list[str] = []

    # ---- Section 1: 标题 + 元信息 ----
    lines.append(f"# 多轮对话评测报告 — {run.name}")
    lines.append("")
    lines.append(f"> Run ID: `#{run.id}` · 状态：**{run.status}**")
    lines.append("")
    if run.description:
        lines.append(f"{run.description}")
        lines.append("")

    lines.append("## 1. 评测元信息")
    lines.append("")
    info_rows = [
        ("评测名称", run.name),
        ("状态", run.status),
        (
            "创建时间",
            run.created_at.strftime("%Y-%m-%d %H:%M:%S") if run.created_at else "-",
        ),
        (
            "开始时间",
            run.started_at.strftime("%Y-%m-%d %H:%M:%S") if run.started_at else "-",
        ),
        (
            "结束时间",
            run.finished_at.strftime("%Y-%m-%d %H:%M:%S") if run.finished_at else "-",
        ),
        ("Dataset", f"#{dataset.id} · {dataset.name}" if dataset else f"#{run.dataset_id}"),
        (
            "Bot 版本",
            f"#{bot.id} · {bot.name} ({bot.version_tag})" if bot else f"#{run.bot_version_id}",
        ),
        (
            "Judge 模型",
            f"#{judge.id} · {judge.name} ({judge.provider}/{judge.model_id})"
            if judge
            else f"#{run.judge_model_id}",
        ),
        ("采样数 / 总数", f"{run.sampling_count or '全量'} / {run.total}"),
        ("已完成 / 失败", f"{run.completed} / {run.failed}"),
        ("加权总分", _md_num(run.weighted_score)),
        ("通过率", _md_pct(run.pass_rate)),
    ]
    lines.append("| 字段 | 值 |")
    lines.append("|------|------|")
    for k, v in info_rows:
        lines.append(f"| {_md_escape(k)} | {_md_escape(v)} |")
    lines.append("")

    # ---- Section 2: 维度汇总 ----
    lines.append("## 2. 维度汇总")
    lines.append("")
    lines.append("| 维度 | 代码 | 权重 | 样本数 | 平均分 | 最低分 | 最高分 | 通过率(≥0.6) |")
    lines.append("|------|------|------|------:|------:|------:|------:|------:|")
    for code in DIM_CODES:
        name = DIMENSION_NAMES.get(code, code)
        weight = DEFAULT_DIMENSION_WEIGHTS.get(code, 0.0)
        row = summary_by_code.get(code, {})
        lines.append(
            "| {name} | `{code}` | {weight} | {n} | {avg} | {mn} | {mx} | {pr} |".format(
                name=_md_escape(name),
                code=code,
                weight=f"{int(weight * 100)}%",
                n=row.get("sample_count", 0),
                avg=_md_num(row.get("avg_score")),
                mn=_md_num(row.get("min_score")),
                mx=_md_num(row.get("max_score")),
                pr=_md_pct(row.get("pass_rate")),
            )
        )
    lines.append("")

    # ---- Section 3: 会话明细（按加权分升序）----
    lines.append("## 3. 会话明细")
    lines.append("")
    sorted_cases = sorted(
        cases,
        key=lambda c: (
            c.weighted_score is None,
            c.weighted_score if c.weighted_score is not None else 0,
        ),
    )
    header = (
        "| case_id | 源会话ID | 加权得分 | "
        + " | ".join(DIMENSION_NAMES.get(code, code) for code in DIM_CODES)
        + " | 最低维度 |"
    )
    sep = "|------:|------|------:|" + ":------:|" * len(DIM_CODES) + "------|"
    lines.append(header)
    lines.append(sep)
    for case in sorted_cases:
        dim_scores = {code: getattr(case, f"{code}_score") for code in DIM_CODES}
        valid = {k: v for k, v in dim_scores.items() if v is not None}
        if valid:
            min_code = min(valid, key=lambda k: valid[k])
            min_label = f"{DIMENSION_NAMES.get(min_code, min_code)} ({_md_num(valid[min_code])})"
        else:
            min_label = "-"
        lines.append(
            "| {cid} | {src} | {w} | {dims} | {minlbl} |".format(
                cid=case.id,
                src=_md_escape(conv_src_map.get(case.conversation_id, f"#{case.conversation_id}")),
                w=_md_num(case.weighted_score),
                dims=" | ".join(_md_num(dim_scores[c]) for c in DIM_CODES),
                minlbl=_md_escape(min_label),
            )
        )
    lines.append("")

    # ---- Section 4: 低分前 5 备注 ----
    lines.append("## 4. 低分 Top-5 备注")
    lines.append("")
    low5 = [c for c in sorted_cases if c.weighted_score is not None][:5]
    if not low5:
        lines.append("_无可用低分样本（所有 case 加权得分均为空）_")
        lines.append("")
    else:
        for c in low5:
            src = conv_src_map.get(c.conversation_id, f"#{c.conversation_id}")
            lines.append(
                f"### case #{c.id} · 源会话 `{src}` · 加权 {_md_num(c.weighted_score)}"
            )
            lines.append("")
            for code in DIM_CODES:
                v = getattr(c, f"{code}_score")
                if v is None:
                    continue
                lines.append(f"- **{DIMENSION_NAMES.get(code, code)}** (`{code}`): {_md_num(v)}")
            note_src = c.error or ""
            if note_src:
                lines.append("")
                note_text = str(note_src)[:300]
                lines.append(f"> 错误：{note_text}")
            lines.append("")

    md_text = "\n".join(lines)
    return md_text.encode("utf-8")


# ============================================================================
# PDF export (C.1) — fpdf2 with latin-1 fallback for CJK
# ============================================================================


def _safe_pdf_text(s: Any) -> str:
    """将任意字符串转为 latin-1 安全的形式（CJK → '?'）。

    fpdf2 默认字体（helvetica）仅支持 latin-1。完整中文请走 MD/XLSX 版本。
    """
    if s is None:
        return "-"
    return str(s).encode("latin-1", "replace").decode("latin-1")


def export_eval_run_pdf(db: Session, eval_run_id: int) -> bytes:
    """生成简易 PDF 报告。非 latin-1 字符（中文等）会被替换为 '?'。"""
    from fpdf import FPDF

    run = db.get(EvalRun, eval_run_id)
    if not run:
        raise ValueError(f"eval run {eval_run_id} not found")

    cases = (
        db.query(EvalCaseResult)
        .filter(EvalCaseResult.eval_run_id == eval_run_id)
        .all()
    )
    dataset = db.get(Dataset, run.dataset_id) if run.dataset_id else None
    bot = db.get(BotVersion, run.bot_version_id) if run.bot_version_id else None
    judge = db.get(JudgeModel, run.judge_model_id) if run.judge_model_id else None

    case_dicts = [
        {f"{code}_score": getattr(c, f"{code}_score") for code in DIM_CODES}
        for c in cases
    ]
    dim_summary = aggregate_dimension_summary(case_dicts, DIM_CODES)
    summary_by_code = {row["dimension_code"]: row for row in dim_summary}

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ---- Disclaimer / Title ----
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(
        0,
        9,
        _safe_pdf_text(f"Eval Run Report #{run.id}: {run.name}"),
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.set_font("helvetica", "I", 8)
    pdf.set_text_color(150, 60, 60)
    pdf.cell(
        0,
        5,
        "Note: non-Latin chars (e.g. Chinese) are replaced with '?'. "
        "For full CJK content, use the MD or XLSX export.",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    # ---- Meta block ----
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(0, 6, "1. Run Metadata", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9)
    meta_rows = [
        ("Status", run.status),
        ("Created", run.created_at.strftime("%Y-%m-%d %H:%M:%S") if run.created_at else "-"),
        ("Started", run.started_at.strftime("%Y-%m-%d %H:%M:%S") if run.started_at else "-"),
        ("Finished", run.finished_at.strftime("%Y-%m-%d %H:%M:%S") if run.finished_at else "-"),
        ("Dataset", f"#{dataset.id} {dataset.name}" if dataset else f"#{run.dataset_id}"),
        (
            "Bot",
            f"#{bot.id} {bot.name} ({bot.version_tag})" if bot else f"#{run.bot_version_id}",
        ),
        (
            "Judge",
            f"#{judge.id} {judge.name} ({judge.provider}/{judge.model_id})"
            if judge
            else f"#{run.judge_model_id}",
        ),
        ("Sampled/Total", f"{run.sampling_count or 'all'} / {run.total}"),
        ("Completed/Failed", f"{run.completed} / {run.failed}"),
        (
            "Weighted Score",
            f"{round(run.weighted_score, 4)}" if run.weighted_score is not None else "-",
        ),
        (
            "Pass Rate",
            f"{run.pass_rate * 100:.1f}%" if run.pass_rate is not None else "-",
        ),
    ]
    for k, v in meta_rows:
        pdf.cell(45, 5, _safe_pdf_text(k))
        pdf.cell(0, 5, _safe_pdf_text(v), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ---- Dimension summary table ----
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(0, 6, "2. Dimension Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "B", 8)
    col_w = [38, 14, 13, 14, 16, 16, 16, 18]
    headers = ["Dimension", "Code", "Weight", "N", "Avg", "Min", "Max", "Pass>=0.6"]
    for w, h in zip(col_w, headers):
        pdf.cell(w, 6, _safe_pdf_text(h), border=1, align="C")
    pdf.ln()

    pdf.set_font("helvetica", "", 8)
    for code in DIM_CODES:
        name = DIMENSION_NAMES.get(code, code)
        weight = DEFAULT_DIMENSION_WEIGHTS.get(code, 0.0)
        row = summary_by_code.get(code, {})
        avg = row.get("avg_score")
        mn = row.get("min_score")
        mx = row.get("max_score")
        pr = row.get("pass_rate")
        n = row.get("sample_count", 0)
        cells = [
            name,
            code,
            f"{int(weight * 100)}%",
            str(n),
            f"{round(avg, 4)}" if avg is not None else "-",
            f"{round(mn, 4)}" if mn is not None else "-",
            f"{round(mx, 4)}" if mx is not None else "-",
            f"{pr * 100:.1f}%" if pr is not None else "-",
        ]
        for w, val in zip(col_w, cells):
            pdf.cell(w, 5, _safe_pdf_text(val), border=1, align="C")
        pdf.ln()
    pdf.ln(3)

    # ---- Low 5 cases ----
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(0, 6, "3. Lowest 5 Cases", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 8)
    low5 = sorted(
        [c for c in cases if c.weighted_score is not None],
        key=lambda c: c.weighted_score,
    )[:5]
    if not low5:
        pdf.cell(0, 5, "(no scored cases)", new_x="LMARGIN", new_y="NEXT")
    else:
        cell_w = [16, 22, 18, 18, 18, 18, 18, 18, 18]
        head = ["case_id", "weighted", "dim1", "dim2", "dim3", "dim4", "dim5", "dim6", "lowest"]
        pdf.set_font("helvetica", "B", 8)
        for w, h in zip(cell_w, head):
            pdf.cell(w, 5, _safe_pdf_text(h), border=1, align="C")
        pdf.ln()
        pdf.set_font("helvetica", "", 8)
        for c in low5:
            cells = [
                str(c.id),
                f"{round(c.weighted_score, 4)}",
            ]
            for code in DIM_CODES:
                v = getattr(c, f"{code}_score")
                cells.append(f"{round(v, 3)}" if v is not None else "-")
            cells.append(c.lowest_dim_code or "-")
            for w, val in zip(cell_w, cells):
                pdf.cell(w, 5, _safe_pdf_text(val), border=1, align="C")
            pdf.ln()

    out = pdf.output()
    # fpdf2 >=2.7 returns bytearray
    return bytes(out)
