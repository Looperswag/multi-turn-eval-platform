"""Dataset 上传 4 步向导通用解析 / 校验 / 转换工具。

工作流：
  1. parse_excel/csv/json(file_bytes) -> ParsedFile  (列 + 扁平行)
  2. infer_field_mapping(columns)     -> FieldMapping （列名启发式）
  3. validate(rows, mapping)          -> ValidationReport
  4. transform(rows, mapping)         -> list[嵌套 conversation payload]

设计原则：
- 内存内处理。任何持久化由调用方在 redis / Postgres 完成。
- 嵌套 JSON（含 turns）会被展平为行；扁平 JSON/CSV/Excel 直接当行处理。
- 启发式列名匹配大小写不敏感，并兼容下划线 / 中划线 / 中文。
"""
from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from typing import Any, Literal

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# 数据类
# ---------------------------------------------------------------------------


@dataclass
class ParsedFile:
    columns: list[str]
    rows: list[dict[str, Any]]
    format: str  # excel / csv / json


TurnIndexSource = Literal["turn_index", "timestamp"]


@dataclass
class FieldMapping:
    conversation_id: str | None = None
    turn_index: str | None = None  # 在 timestamp 模式下，这里也存"时间戳列名"
    user_query: str | None = None
    rewritten_query: str | None = None
    dimension_tag: str | None = None
    quality_label: str | None = None
    issue_type: str | None = None
    # 轮次序号的派生方式：
    #   "turn_index" — 直接读取数字列（默认）
    #   "timestamp" — 把 turn_index 字段视作时间戳列，按 conv 内升序派生 1..N
    turn_index_source: TurnIndexSource = "turn_index"

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class ValidationIssue:
    severity: str  # error / warning / info
    code: str
    message: str
    count: int
    sample: list[str] = field(default_factory=list)


@dataclass
class ValidationReport:
    issues: list[ValidationIssue]
    total_conversations: int
    total_turns: int
    is_passable: bool


# ---------------------------------------------------------------------------
# 解析
# ---------------------------------------------------------------------------


def parse_excel(file_bytes: bytes, sheet_name: int | str = 0) -> ParsedFile:
    """读 .xlsx 首 sheet（或指定 sheet），第一行作 header。空值统一转 ""。"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return ParsedFile(columns=[], rows=[], format="excel")

    columns = [
        str(c).strip() if c is not None else f"col_{i + 1}"
        for i, c in enumerate(header_row)
    ]

    rows: list[dict[str, Any]] = []
    for raw in row_iter:
        if raw is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        row = {columns[i]: ("" if raw[i] is None else raw[i]) for i in range(min(len(columns), len(raw)))}
        rows.append(row)

    return ParsedFile(columns=columns, rows=rows, format="excel")


def parse_csv(file_bytes: bytes) -> ParsedFile:
    """用 csv.DictReader 读 CSV；尝试 utf-8 / utf-8-sig / gbk 三种编码。"""
    text: str | None = None
    last_err: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError as e:
            last_err = e
    if text is None:
        raise ValueError(f"unable to decode csv: {last_err}")

    reader = csv.DictReader(io.StringIO(text))
    columns = [c.strip() for c in (reader.fieldnames or [])]
    rows = [dict(r) for r in reader]
    return ParsedFile(columns=columns, rows=rows, format="csv")


def parse_json(file_bytes: bytes) -> ParsedFile:
    """JSON 支持两种格式：
    1. 嵌套（mock_multi_turn_queries_100.json 同款）：[{conversation_id, turns: [...]}]
       -> 展平为行：每个 turn 一行，附带 conversation 级元信息
    2. 扁平：[{conversation_id, turn_index, user_query, ...}]
    """
    text = file_bytes.decode("utf-8-sig", errors="replace")
    data = json.loads(text)
    if not isinstance(data, list):
        raise ValueError("expect json array at top level")

    if not data:
        return ParsedFile(columns=[], rows=[], format="json")

    # 判定嵌套：第一条含 turns list
    first = data[0]
    if isinstance(first, dict) and isinstance(first.get("turns"), list):
        rows: list[dict[str, Any]] = []
        for conv in data:
            conv_id = conv.get("conversation_id")
            for t in conv.get("turns", []):
                row = {
                    "conversation_id": conv_id,
                    "dimension_tag": conv.get("dimension_tag"),
                    "quality_label": conv.get("quality_label"),
                    "issue_type": conv.get("issue_type"),
                    "turn_index": t.get("turn_index"),
                    "user_query": t.get("user_query"),
                    "rewritten_query": t.get("rewritten_query"),
                    "timestamp": t.get("timestamp"),
                    "query_id": t.get("query_id"),
                }
                rows.append(row)
        cols = list(rows[0].keys()) if rows else [
            "conversation_id", "dimension_tag", "quality_label", "issue_type",
            "turn_index", "user_query", "rewritten_query", "timestamp", "query_id",
        ]
        return ParsedFile(columns=cols, rows=rows, format="json")

    # 扁平：用并集列名
    cols_set: list[str] = []
    seen: set[str] = set()
    for r in data:
        if not isinstance(r, dict):
            continue
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                cols_set.append(k)
    return ParsedFile(columns=cols_set, rows=[dict(r) for r in data], format="json")


# ---------------------------------------------------------------------------
# 启发式映射
# ---------------------------------------------------------------------------


# 关键字 -> 目标字段。匹配规则：lower-strip + 去掉下划线/中划线/空格 后整词或子串匹配。
_FIELD_ALIASES: dict[str, list[str]] = {
    "conversation_id": [
        "conversation_id", "conv_id", "conversationid", "convid",
        "session_id", "sessionid", "dialogue_id", "dialog_id",
        # 线上简单导出格式：meta_id 是 session 维度的标识（同 meta_id = 同一会话）；
        # 优先级前置以避免被 "id" 列名经子串匹配抢走。
        "meta_id", "metaid", "meta_conversation_id",
        "会话id", "对话id", "session", "conversation",
    ],
    "turn_index": [
        "turn_index", "turn", "turnidx", "turnindex", "turnno", "turn_no",
        "round", "round_index", "step", "step_index",
        "轮次", "轮序", "轮数", "turn轮次", "顺序",
    ],
    # 仅用于启发式：识别后将作为 turn_index 列 + turn_index_source="timestamp"
    "_timestamp": [
        "gmt_create", "gmtcreate", "created_at", "createdat", "create_time",
        "createtime", "timestamp", "time", "datetime", "event_time", "ts",
        "创建时间", "时间", "时间戳",
    ],
    "user_query": [
        "user_query", "userquery", "query", "user_input", "userinput",
        "input", "utterance", "user", "user_text", "raw_query", "原始query",
        "用户输入", "用户问题", "问题", "用户query",
    ],
    "rewritten_query": [
        "rewritten_query", "rewrittenquery", "rewrite", "rewrite_query",
        "rewritten", "bot_rewrite", "bot_query", "expanded_query",
        "改写query", "改写", "改写后", "重写query",
    ],
    "dimension_tag": [
        "dimension_tag", "dimension", "dim", "dim_tag", "category",
        "维度", "维度标签", "标签",
    ],
    "quality_label": [
        "quality_label", "quality", "label", "good_bad", "is_good",
        "质量", "质量标签", "正负样本", "good/bad",
    ],
    "issue_type": [
        "issue_type", "issue", "issuetype", "bug_type", "error_type",
        "问题类型", "问题",
    ],
}


def _normalize(col: str) -> str:
    return col.strip().lower().replace("_", "").replace("-", "").replace(" ", "")


def infer_field_mapping(columns: list[str]) -> FieldMapping:
    """按列名启发式推断映射。每个目标字段只匹配第一个命中的列。

    `_timestamp` 不是真实字段，仅作为 turn_index 列的备选：当没识别到数字 turn_index 列
    但识别到时间戳列时，自动把 mapping.turn_index 指向时间戳列并切到 timestamp 模式。
    """
    normalized = [(c, _normalize(c)) for c in columns]
    mapping = FieldMapping()
    used_cols: set[str] = set()
    inferred: dict[str, str] = {}  # target -> chosen col

    for target, aliases in _FIELD_ALIASES.items():
        norm_aliases = [_normalize(a) for a in aliases]
        chosen: str | None = None
        for col, n in normalized:
            if col in used_cols:
                continue
            if n in norm_aliases:
                chosen = col
                break
        if chosen is None:
            for col, n in normalized:
                if col in used_cols:
                    continue
                for a in norm_aliases:
                    if not a:
                        continue
                    if a in n or n in a:
                        chosen = col
                        break
                if chosen:
                    break
        if chosen:
            inferred[target] = chosen
            used_cols.add(chosen)

    # 应用真实字段
    for target, col in inferred.items():
        if target.startswith("_"):
            continue
        setattr(mapping, target, col)

    # 若没识别到 turn_index 但识别到 timestamp 列，按时间戳派生
    if not mapping.turn_index and inferred.get("_timestamp"):
        mapping.turn_index = inferred["_timestamp"]
        mapping.turn_index_source = "timestamp"

    return mapping


# ---------------------------------------------------------------------------
# 校验
# ---------------------------------------------------------------------------


_REQUIRED_FIELDS = ("conversation_id", "turn_index", "user_query")
_MAX_SAMPLE = 5


def _to_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


# 常见时间戳格式（按宽容程度从严到松；首个能解析的胜出）
_DATETIME_FORMATS: tuple[str, ...] = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%SZ",
    "%Y-%m-%d",
    "%Y/%m/%d",
)


def _parse_datetime(v: Any) -> datetime | None:
    """尽力把任意值解析为 datetime。

    支持：
    - datetime / date 对象（直接转换）
    - int/float 当作 epoch 秒；> 1e12 时按毫秒处理
    - str：先尝试 fromisoformat，再走若干常见格式
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    # date 但不是 datetime
    try:
        from datetime import date
        if isinstance(v, date):
            return datetime(v.year, v.month, v.day)
    except Exception:
        pass

    if isinstance(v, (int, float)):
        try:
            ts = float(v)
            if ts > 1e12:  # 毫秒级
                ts = ts / 1000.0
            return datetime.fromtimestamp(ts, tz=timezone.utc).replace(tzinfo=None)
        except (OverflowError, OSError, ValueError):
            return None

    s = str(v).strip()
    if not s:
        return None
    # 纯数字字符串当 epoch
    if s.isdigit():
        try:
            ts = float(s)
            if ts > 1e12:
                ts = ts / 1000.0
            return datetime.fromtimestamp(ts, tz=timezone.utc).replace(tzinfo=None)
        except (OverflowError, OSError, ValueError):
            pass
    # ISO 8601（Python 3.11+ fromisoformat 容忍 'Z' 后缀，但更早版本不行 → 简单兜底）
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        pass
    for fmt in _DATETIME_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def validate(rows: list[dict[str, Any]], mapping: FieldMapping) -> ValidationReport:
    issues: list[ValidationIssue] = []

    # 1. 检查必填字段映射存在
    missing_mappings = [f for f in _REQUIRED_FIELDS if not getattr(mapping, f)]
    if missing_mappings:
        issues.append(
            ValidationIssue(
                severity="error",
                code="missing_field_mapping",
                message=f"必填字段未映射：{', '.join(missing_mappings)}",
                count=len(missing_mappings),
                sample=missing_mappings,
            )
        )
        # 直接返回，后续校验依赖映射
        return ValidationReport(issues=issues, total_conversations=0, total_turns=0, is_passable=False)

    cid_col = mapping.conversation_id
    tidx_col = mapping.turn_index
    uq_col = mapping.user_query
    by_timestamp = mapping.turn_index_source == "timestamp"

    empty_conv: list[str] = []
    empty_query: list[str] = []
    invalid_turn_idx: list[str] = []
    invalid_timestamps: list[str] = []

    # 按 conversation_id 聚合
    conv_turns: dict[str, list[int]] = {}          # 仅 turn_index 模式用
    conv_timestamps: dict[str, list[datetime]] = {} # 仅 timestamp 模式用
    conv_seen_pairs: dict[tuple[str, Any], int] = {}
    dup_pairs: list[str] = []

    for i, row in enumerate(rows):
        conv_id = _to_str(row.get(cid_col))
        uq = _to_str(row.get(uq_col))

        if not conv_id:
            if len(empty_conv) < _MAX_SAMPLE:
                empty_conv.append(f"row#{i + 2}")
            continue

        if by_timestamp:
            ts = _parse_datetime(row.get(tidx_col))
            if ts is None:
                if len(invalid_timestamps) < _MAX_SAMPLE:
                    invalid_timestamps.append(f"row#{i + 2} value={row.get(tidx_col)!r}")
                continue
            key: tuple[str, Any] = (conv_id, ts)
            if key in conv_seen_pairs:
                if len(dup_pairs) < _MAX_SAMPLE:
                    dup_pairs.append(f"{conv_id}@{ts.isoformat()}")
                continue
            conv_seen_pairs[key] = i
            conv_timestamps.setdefault(conv_id, []).append(ts)
            if not uq and len(empty_query) < _MAX_SAMPLE:
                empty_query.append(f"{conv_id}@{ts.isoformat()}")
        else:
            tidx = _to_int(row.get(tidx_col))
            if tidx is None:
                if len(invalid_turn_idx) < _MAX_SAMPLE:
                    invalid_turn_idx.append(f"row#{i + 2} value={row.get(tidx_col)!r}")
                continue
            if not uq:
                if len(empty_query) < _MAX_SAMPLE:
                    empty_query.append(f"{conv_id}/turn#{tidx}")
            key = (conv_id, tidx)
            if key in conv_seen_pairs:
                if len(dup_pairs) < _MAX_SAMPLE:
                    dup_pairs.append(f"{conv_id}/turn#{tidx}")
                continue
            conv_seen_pairs[key] = i
            conv_turns.setdefault(conv_id, []).append(tidx)

    if empty_conv:
        issues.append(ValidationIssue(
            severity="error", code="empty_conversation_id",
            message="存在空 conversation_id 的行", count=len(empty_conv), sample=empty_conv,
        ))
    if invalid_turn_idx:
        issues.append(ValidationIssue(
            severity="error", code="invalid_turn_index",
            message="turn_index 无法解析为整数", count=len(invalid_turn_idx), sample=invalid_turn_idx,
        ))
    if invalid_timestamps:
        issues.append(ValidationIssue(
            severity="error", code="invalid_timestamp",
            message="时间戳无法解析（支持 YYYY-MM-DD HH:MM:SS、ISO 8601、epoch 等）",
            count=len(invalid_timestamps), sample=invalid_timestamps,
        ))
    if empty_query:
        issues.append(ValidationIssue(
            severity="error", code="empty_user_query",
            message="user_query 为空", count=len(empty_query), sample=empty_query,
        ))
    if dup_pairs:
        issues.append(ValidationIssue(
            severity="warning",
            code="duplicate_pair",
            message=(
                "同一 conversation 内出现重复时间戳，将按出现顺序保留首条"
                if by_timestamp
                else "重复 (conversation_id, turn_index) 已被忽略"
            ),
            count=len(dup_pairs), sample=dup_pairs,
        ))

    long_convs: list[str] = []
    if by_timestamp:
        # timestamp 模式：派生后必然 1..N 连续，不需要 gap 校验；
        # 但若同一会话只有 1 轮，未必是问题，跳过。
        for cid, tss in conv_timestamps.items():
            if len(tss) > 20 and len(long_convs) < _MAX_SAMPLE:
                long_convs.append(f"{cid} ({len(tss)} turns)")
    else:
        turn_gaps: list[str] = []
        for cid, idxs in conv_turns.items():
            idxs_sorted = sorted(idxs)
            if idxs_sorted[0] != 1 or any(b - a != 1 for a, b in zip(idxs_sorted, idxs_sorted[1:])):
                if len(turn_gaps) < _MAX_SAMPLE:
                    turn_gaps.append(f"{cid} turns={idxs_sorted}")
            if len(idxs_sorted) > 20 and len(long_convs) < _MAX_SAMPLE:
                long_convs.append(f"{cid} ({len(idxs_sorted)} turns)")
        if turn_gaps:
            issues.append(ValidationIssue(
                severity="warning", code="turn_index_gap",
                message="同一 conversation 的 turn_index 不连续或未从 1 开始",
                count=len(turn_gaps), sample=turn_gaps,
            ))
    if long_convs:
        issues.append(ValidationIssue(
            severity="warning", code="overlong_conversation",
            message="单个 conversation 轮次过多（> 20）",
            count=len(long_convs), sample=long_convs,
        ))

    # info：首轮无 rewritten_query 属正常
    if mapping.rewritten_query:
        first_no_rewrite = 0
        for row in rows:
            if _to_int(row.get(tidx_col)) == 1:
                rw = _to_str(row.get(mapping.rewritten_query))
                if not rw:
                    first_no_rewrite += 1
        if first_no_rewrite > 0:
            issues.append(ValidationIssue(
                severity="info", code="first_turn_no_rewrite",
                message="首轮 rewritten_query 为空（正常）",
                count=first_no_rewrite, sample=[],
            ))

    is_passable = not any(i.severity == "error" for i in issues)
    total_convs = len(conv_timestamps) if by_timestamp else len(conv_turns)
    return ValidationReport(
        issues=issues,
        total_conversations=total_convs,
        total_turns=len(conv_seen_pairs),
        is_passable=is_passable,
    )


# ---------------------------------------------------------------------------
# 转换为入库格式
# ---------------------------------------------------------------------------


def transform(rows: list[dict[str, Any]], mapping: FieldMapping) -> list[dict[str, Any]]:
    """按 (conversation_id) 分组拼装为现有 datasets.upload 接受的嵌套格式。

    turn_index_source:
      - "turn_index": 直接读列值为整数；duplicate (cid, tidx) 后写入者被丢弃。
      - "timestamp" : 列值解析为 datetime；每会话按时间升序派生 turn_index 1..N；
                      原始时间戳字符串写入 Turn.timestamp。
    """
    if not all(getattr(mapping, f) for f in _REQUIRED_FIELDS):
        raise ValueError("mapping missing required fields")

    cid_col = mapping.conversation_id
    tidx_col = mapping.turn_index
    uq_col = mapping.user_query
    rw_col = mapping.rewritten_query
    dim_col = mapping.dimension_tag
    ql_col = mapping.quality_label
    iss_col = mapping.issue_type
    by_timestamp = mapping.turn_index_source == "timestamp"

    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        cid = _to_str(row.get(cid_col))
        if not cid:
            continue
        uq = _to_str(row.get(uq_col))
        if not uq:
            continue

        if by_timestamp:
            ts_obj = _parse_datetime(row.get(tidx_col))
            if ts_obj is None:
                continue
            sort_key: Any = ts_obj
            # 保留原始时间戳的字符串形态（去掉 ws）作为 Turn.timestamp
            ts_display = _to_str(row.get(tidx_col)) or ts_obj.strftime("%Y-%m-%d %H:%M:%S")
        else:
            tidx_val = _to_int(row.get(tidx_col))
            if tidx_val is None:
                continue
            sort_key = tidx_val
            ts_display = _to_str(row.get("timestamp")) or None

        conv = grouped.get(cid)
        if conv is None:
            conv = {
                "conversation_id": cid,
                "dimension_tag": _to_str(row.get(dim_col)) if dim_col else None,
                "quality_label": _to_str(row.get(ql_col)) if ql_col else None,
                "issue_type": _to_str(row.get(iss_col)) if iss_col else None,
                "_pending_turns": [],
                "_seen_keys": set(),
            }
            for k in ("dimension_tag", "quality_label", "issue_type"):
                if conv[k] == "":
                    conv[k] = None
            grouped[cid] = conv

        if sort_key in conv["_seen_keys"]:
            continue
        conv["_seen_keys"].add(sort_key)

        pending = {
            "_sort_key": sort_key,
            "user_query": uq,
        }
        if rw_col:
            rw = _to_str(row.get(rw_col))
            # Hive/SQL 导出常用 \N 表示 NULL；NULL/none 之类的字面 NULL 标记也统一归 None。
            # 否则首轮真实的"无改写"会以字面 "\N" 残留进 BotRewrite，导致后续 evaluator
            # 把它当作字符串"改写"，破坏 dim1 的「首轮无改写跳过」语义。
            if rw.upper() in (r"\N", "NULL", "NONE"):
                rw = ""
            pending["rewritten_query"] = rw or None
        if ts_display:
            pending["timestamp"] = ts_display
        conv["_pending_turns"].append(pending)

    out: list[dict[str, Any]] = []
    for conv in grouped.values():
        pending = sorted(conv.pop("_pending_turns"), key=lambda t: t["_sort_key"])
        turns: list[dict[str, Any]] = []
        for i, p in enumerate(pending):
            t = {k: v for k, v in p.items() if k != "_sort_key"}
            # timestamp 模式下派生 1..N；turn_index 模式下保留原值
            t["turn_index"] = (i + 1) if by_timestamp else p["_sort_key"]
            turns.append(t)
        conv["turns"] = turns
        conv["total_turns"] = len(turns)
        conv.pop("_seen_keys", None)
        out.append(conv)
    return out


# ---------------------------------------------------------------------------
# 顶层入口
# ---------------------------------------------------------------------------


def parse_any(file_bytes: bytes, filename: str, format_hint: str | None = None) -> ParsedFile:
    """按文件扩展名 / format hint 调度到具体 parser。"""
    fmt = (format_hint or "").lower().strip()
    name = (filename or "").lower()
    if fmt:
        if fmt in ("excel", "xlsx"):
            return parse_excel(file_bytes)
        if fmt == "csv":
            return parse_csv(file_bytes)
        if fmt == "json":
            return parse_json(file_bytes)
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return parse_excel(file_bytes)
    if name.endswith(".csv"):
        return parse_csv(file_bytes)
    if name.endswith(".json"):
        return parse_json(file_bytes)
    # 兜底尝试 JSON
    return parse_json(file_bytes)


# ---------------------------------------------------------------------------
# 线上格式（A.4）：来自生产抓取的 Excel
#   meta_conversation_id + historyquery + llm_resp 三列同时存在即触发
# 直接产出嵌套 conv list（与 transform() 输出兼容），不走通用启发式映射
# ---------------------------------------------------------------------------


ONLINE_REQUIRED_COLUMNS = ("meta_conversation_id", "historyquery", "llm_resp")
# 「线上简单格式」：生产线另一种导出 schema — 一行一轮，
# 由 meta_id 分组、gmt_create 排序、首轮 rewritten_query 通常为 \N。
# 命中后短路 4 步向导，沿用既有 transform() 完成嵌套化。
ONLINE_SIMPLE_REQUIRED_COLUMNS = ("meta_id", "gmt_create", "user_query", "rewritten_query")
# 解析 historyquery 字段中 `#第N轮问题：...` 与 `#第N轮追问：...` 段
# 用 lookahead 把 (.+?) 截到下一个 `#第` 或字符串末尾，DOTALL 让 `.` 跨行匹配
_HISTORY_BLOCK_RE = re.compile(
    r"#第\s*(\d+)\s*轮\s*(问题|追问)\s*[：:]\s*(.*?)(?=\n\s*#第\s*\d+\s*轮|\Z)",
    re.DOTALL,
)


def is_online_format(columns: list[str]) -> bool:
    """三列同时出现才算命中（严格判定，避免误伤其它 Excel）。"""
    lower = {(c or "").strip().lower() for c in columns}
    return all(c.lower() in lower for c in ONLINE_REQUIRED_COLUMNS)


def is_online_simple_format(columns: list[str]) -> bool:
    """meta_id + gmt_create + user_query + rewritten_query 四列同时出现即命中。"""
    lower = {(c or "").strip().lower() for c in columns}
    return all(c.lower() in lower for c in ONLINE_SIMPLE_REQUIRED_COLUMNS)


def parse_online_simple_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """线上简单格式 Excel → 嵌套 conv list（与 parse_online_excel 输出结构兼容）。

    内部直接复用 parse_excel + transform，传固定 FieldMapping（meta_id→conv_id、
    gmt_create→turn_index_source=timestamp）。这样向导可直接进入预览/入库。
    """
    parsed = parse_excel(file_bytes)
    mapping = FieldMapping(
        conversation_id="meta_id",
        turn_index="gmt_create",
        user_query="user_query",
        rewritten_query="rewritten_query",
        turn_index_source="timestamp",
    )
    # 同列名可能存在大小写差异；规一化到 mapping 期望的字面值
    norm_to_col = {(c or "").strip().lower(): c for c in parsed.columns}
    for fld in ("conversation_id", "turn_index", "user_query", "rewritten_query"):
        target = getattr(mapping, fld)
        if target and target.lower() in norm_to_col:
            setattr(mapping, fld, norm_to_col[target.lower()])
    return transform(parsed.rows, mapping)


def _parse_history_query(history_text: str) -> list[dict[str, Any]]:
    """从 historyquery 字段抽出每轮 `(turn_index, kind, text)`，按 turn_index 聚合。

    返回 list[{"turn_index": int, "user_query": str | None, "bot_response": str | None}]，
    按 turn_index 升序。同 turn 出现多次问题/追问时拼接为多行字符串。
    """
    if not history_text:
        return []
    by_turn: dict[int, dict[str, list[str]]] = {}
    for m in _HISTORY_BLOCK_RE.finditer(history_text):
        idx = int(m.group(1))
        kind = m.group(2)
        text = m.group(3).strip()
        if not text:
            continue
        slot = by_turn.setdefault(idx, {"问题": [], "追问": []})
        slot[kind].append(text)
    out: list[dict[str, Any]] = []
    for idx in sorted(by_turn.keys()):
        s = by_turn[idx]
        out.append({
            "turn_index": idx,
            "user_query": ("\n".join(s["问题"]) if s["问题"] else None),
            "bot_response": ("\n".join(s["追问"]) if s["追问"] else None),
        })
    return out


def _parse_llm_resp(raw: Any) -> dict[str, Any]:
    """`llm_resp` 列既可能是 dict、也可能是 JSON 字符串。返回标准化 dict。
    任何解析失败均回空 dict，不抛错（数据脏的可能性较高）。
    """
    if raw is None or raw == "":
        return {}
    if isinstance(raw, dict):
        d = raw
    elif isinstance(raw, str):
        try:
            d = json.loads(raw)
        except (json.JSONDecodeError, ValueError):
            return {}
    else:
        return {}
    if not isinstance(d, dict):
        return {}
    return d


def _normalize_constraints(val: Any) -> list[str] | None:
    """约束列表标准化：None / 空 list / 非 list 都归一为 None，否则保留 str 元素。"""
    if not isinstance(val, list):
        return None
    items = [str(x).strip() for x in val if x is not None and str(x).strip()]
    return items if items else None


def parse_online_excel(file_bytes: bytes, sheet_name: int | str = 0) -> list[dict[str, Any]]:
    """线上 Excel 直接产出嵌套 conv list（每条 conv 含 turns，turn 含 bot 元信息）。

    规则：
    1. 按 `meta_conversation_id` 分组
    2. 同 conv 内按 `gmt_create` 升序去重；同 `(conv_id, turn_index_from_history)` 取
       gmt_create 最大者
    3. 用 `historyquery` 重构 phantom 前置轮次（user_query 来自 `#第N轮问题`，
       bot_response 来自 `#第N轮追问`，rewritten_query=None，其余 bot 元字段=None）
    4. 当前 Excel 行 = phantom 列表之后的下一轮：turn_index = max(phantom) + 1
    5. 输出 turn dict：
        {
          "turn_index", "user_query", "rewritten_query", "timestamp",
          "bot_response", "intent_type",
          "inherited_constraints", "dropped_constraints", "needs_rewrite"
        }
    """
    parsed = parse_excel(file_bytes, sheet_name=sheet_name)
    rows = parsed.rows
    if not rows:
        return []

    # 用列名小写化简化访问
    def col(row: dict, name: str) -> Any:
        # 兼容大小写
        if name in row:
            return row[name]
        for k in row.keys():
            if k.lower() == name.lower():
                return row[k]
        return None

    # 按 conv_id 分组
    by_conv: dict[str, list[dict]] = {}
    for r in rows:
        cid_raw = col(r, "meta_conversation_id")
        cid = "" if cid_raw is None else str(cid_raw).strip()
        if not cid:
            continue
        by_conv.setdefault(cid, []).append(r)

    out: list[dict[str, Any]] = []
    for cid, raw_rows in by_conv.items():
        # 行级排序：按 gmt_create 升序；同 ts 去重保留首条
        def _ts(r: dict) -> datetime | None:
            return _parse_datetime(col(r, "gmt_create"))
        sorted_rows = sorted(
            raw_rows,
            key=lambda r: (_ts(r) or datetime.min),
        )

        # 用 historyquery 最长的那条作为前置基线（避免短 history 覆盖长 history）
        baseline_row = max(
            sorted_rows,
            key=lambda r: len(str(col(r, "historyquery") or "")),
        )
        phantom_turns = _parse_history_query(str(col(baseline_row, "historyquery") or ""))

        # turn_index → turn dict
        merged: dict[int, dict[str, Any]] = {}
        for ph in phantom_turns:
            merged[ph["turn_index"]] = {
                "turn_index": ph["turn_index"],
                "user_query": ph["user_query"] or "",
                "rewritten_query": None,
                "timestamp": None,
                "bot_response": ph["bot_response"],
                "intent_type": None,
                "inherited_constraints": None,
                "dropped_constraints": None,
                "needs_rewrite": None,
            }

        # 把 Excel 各行作为"新轮"叠加（覆盖同 idx 的 phantom）
        # 关键：每行的 turn_index = 该行 historyquery 已含轮次的最大值 + 1
        # 若没有 historyquery 则用全局推导（按 ts 升序 1..N）
        ts_sorted_rows = [(r, _ts(r)) for r in sorted_rows]
        fallback_idx = 0
        for row, ts in ts_sorted_rows:
            history = _parse_history_query(str(col(row, "historyquery") or ""))
            if history:
                this_idx = max(t["turn_index"] for t in history) + 1
            else:
                fallback_idx += 1
                this_idx = fallback_idx
            user_q = str(col(row, "ori_query") or col(row, "query") or "").strip()
            rewrite = str(col(row, "rewritten_query") or "").strip() or None
            meta = _parse_llm_resp(col(row, "llm_resp"))
            ts_str = (
                ts.strftime("%Y-%m-%d %H:%M:%S")
                if isinstance(ts, datetime)
                else (str(col(row, "gmt_create")).strip() or None)
            )
            # 该行所在 turn 的 bot_response 优先从 history 里取（若 history 有这一 idx 的追问），
            # 否则保留 None（这是"当前正在 evaluate 的轮"，bot 还没回 → null 合理）
            bot_resp_from_history = next(
                (t["bot_response"] for t in history if t["turn_index"] == this_idx),
                None,
            )
            merged[this_idx] = {
                "turn_index": this_idx,
                "user_query": user_q,
                "rewritten_query": rewrite,
                "timestamp": ts_str,
                "bot_response": bot_resp_from_history,
                "intent_type": (meta.get("intent_type") or None),
                "inherited_constraints": _normalize_constraints(meta.get("inherited_constraints")),
                "dropped_constraints": _normalize_constraints(meta.get("dropped_constraints")),
                "needs_rewrite": (
                    bool(meta["needs_rewrite"]) if "needs_rewrite" in meta else None
                ),
            }

        # 按 idx 升序
        turns_sorted = [merged[i] for i in sorted(merged.keys()) if merged[i]["user_query"]]
        if not turns_sorted:
            continue
        # 重新归一化 turn_index 为 1..N（避免 historyquery 跳号导致评测端断序）
        for new_idx, t in enumerate(turns_sorted, start=1):
            t["turn_index"] = new_idx

        out.append({
            "conversation_id": cid,
            "dimension_tag": None,
            "quality_label": None,
            "issue_type": None,
            "total_turns": len(turns_sorted),
            "turns": turns_sorted,
        })

    # 按 conversation_id 排序保持稳定性
    out.sort(key=lambda c: c["conversation_id"])
    return out
