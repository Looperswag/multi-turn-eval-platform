"""导出器 5-sheet 升级（A.4）端到端单测。

在真实 DB（sqlite 内存或 docker postgres）里搭一个最小评测 run，然后调
`export_eval_run_xlsx`，断言：
- Sheet "Per-Turn 维度明细" 每个 (case, turn) 一行，含 user_query/rewritten_query/bot_response/intent_type/reasoning
- Sheet "Session 维度明细" 每条 case 一行，含 dim2 false_inherited/missed_constraints/precision/recall
- "总览"、"维度汇总"、"会话明细" 三个老 sheet 仍在
"""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from app.core.db import SessionLocal
from app.models import (
    BotRewrite,
    BotVersion,
    Conversation,
    Dataset,
    EvalCaseResult,
    EvalRun,
    EvalTurnResult,
    JudgeModel,
    JudgePromptVersion,
    Turn,
)
from app.services.exporter import export_eval_run_xlsx


# ---------------------------------------------------------------------------
# Fixture：搭最小测试 run
# ---------------------------------------------------------------------------

def _cleanup_test_data(db, tag: str) -> None:
    """按外键依赖正序删：EvalRun → Dataset → BotVersion。
    每一步通过 join 找到所有引用我们 tag 的下游记录，避免遗漏。"""
    # 1. EvalRun：本测的 bot_version / dataset 都标了 tag；找所有引用这些的 run
    bot_ids = [b.id for b in db.query(BotVersion).filter(BotVersion.version_tag == tag).all()]
    ds_ids = [d.id for d in db.query(Dataset).filter(Dataset.name.like(f"{tag}%")).all()]
    if bot_ids or ds_ids:
        runs_to_del = db.query(EvalRun).filter(
            (EvalRun.bot_version_id.in_(bot_ids) if bot_ids else False)
            | (EvalRun.dataset_id.in_(ds_ids) if ds_ids else False)
        ).all()
        for r in runs_to_del:
            db.delete(r)
        db.flush()
    # 2. Dataset（Conversation/Turn 走 CASCADE）
    for d in db.query(Dataset).filter(Dataset.name.like(f"{tag}%")).all():
        db.delete(d)
    db.flush()
    # 3. BotVersion（最后；BotRewrite → Turn → CASCADE 已带走，所以 BotVersion 可独立删）
    for b in db.query(BotVersion).filter(BotVersion.version_tag == tag).all():
        db.delete(b)
    db.commit()


@pytest.fixture
def synthetic_run():
    db = SessionLocal()
    try:
        tag = "v5export-test"
        _cleanup_test_data(db, tag)

        bot = BotVersion(name=f"{tag}-bot", version_tag=tag, description="test")
        dataset = Dataset(name=f"{tag}-ds", description="test", version="v1",
                          conversation_count=1)
        judge = (
            db.query(JudgeModel).first()
            or JudgeModel(name="dummy", provider="deepseek", model_id="dummy")
        )
        if judge.id is None:
            db.add(judge)
        db.add_all([bot, dataset])
        db.flush()

        # 一条会话，3 轮，phantom turn 1
        conv = Conversation(
            dataset_id=dataset.id,
            conversation_id_src=f"{tag}_conv1",
            total_turns=3,
        )
        db.add(conv)
        db.flush()

        turns_data = [
            (1, "帮我找优惠", None, None, None, None),
            (2, "本人身形偏瘦，帮我搭配夏天短袖",
             "推荐适合身形偏瘦的夏季短袖",
             "你对哪款比较感兴趣？",
             "商品检索",
             ["身形偏瘦", "夏季短袖"]),
            (3, "多推几款",
             "请推荐几款适合身形偏瘦的夏季短袖",
             None,
             "商品检索",
             ["夏季短袖", "身形偏瘦"]),
        ]
        for tidx, uq, rq, br, intent, inh in turns_data:
            t = Turn(conversation_id=conv.id, turn_index=tidx, user_query=uq)
            db.add(t)
            db.flush()
            db.add(BotRewrite(
                turn_id=t.id,
                bot_version_id=bot.id,
                rewritten_query=rq,
                bot_response=br,
                intent_type=intent,
                inherited_constraints=inh,
                dropped_constraints=[],
                needs_rewrite=(rq is not None),
            ))
        db.flush()

        # prompt 版本（任选活跃版即可）
        prompt_v5 = db.query(JudgePromptVersion).filter(
            JudgePromptVersion.version_tag == "v5",
            JudgePromptVersion.dimension_code == "dim1",
        ).first()
        prompt_v5_id = prompt_v5.id if prompt_v5 else None

        run = EvalRun(
            name=f"{tag}-run",
            description="A.4 exporter test",
            status="success",
            dataset_id=dataset.id,
            bot_version_id=bot.id,
            judge_model_id=judge.id,
            judge_prompt_version_ids={"dim1": prompt_v5_id} if prompt_v5_id else {},
            dimensions_selected=["dim1", "dim2", "dim3", "dim4", "dim5", "dim6"],
            total=1,
            completed=1,
            failed=0,
            weighted_score=0.85,
            pass_rate=1.0,
        )
        db.add(run)
        db.flush()

        case = EvalCaseResult(
            eval_run_id=run.id,
            conversation_id=conv.id,
            weighted_score=0.85,
            lowest_dim_code="dim4",
            dim1_score=0.9, dim2_score=0.8, dim3_score=1.0,
            dim4_score=0.5, dim5_score=1.0, dim6_score=None,
            dim_results_full={
                "dim1": {"score": 0.9, "detail": {"overall_score": 0.9, "explanation": "改写忠实"}},
                "dim2": {"score": 0.8, "detail": {
                    "false_inherited": [
                        {"turn": 2, "claimed_constraint": "身高1.71米",
                         "evidence_in_rewrite": "未出现"}
                    ],
                    "missed_constraints": [],
                    "correctly_inherited_count": 3,
                    "precision": 0.75,
                    "recall": 1.0,
                    "overall_score": 0.875,
                    "explanation": "dim2 一条 false_inherited",
                }},
                "dim3": {"score": 1.0, "detail": {"applicable": False, "score": 1.0}},
                "dim4": {"score": 0.5, "detail": {"applicable": True, "score": 0.5,
                                                   "anaphora_type": "single",
                                                   "explanation": "部分消解"}},
                "dim5": {"score": 1.0, "detail": {"applicable": True, "score": 1.0}},
                "dim6": {"score": None, "detail": {"applicable": False,
                                                    "explanation": "未识别到纠错"}},
            },
        )
        db.add(case)
        db.flush()

        # turn-level dim 结果
        turn_evals = [
            (2, "dim1", 0.9, None, {"overall_score": 0.9, "explanation": "turn2 dim1 ok"}),
            (3, "dim1", 0.9, None, {"overall_score": 0.9, "explanation": "turn3 dim1 ok"}),
            (2, "dim3", None, False, {"applicable": False, "boundary_type": "normal_shopping",
                                       "score": 0, "explanation": "正常导购"}),
            (3, "dim3", None, False, {"applicable": False, "boundary_type": "normal_shopping",
                                       "score": 0, "explanation": "正常导购"}),
            (3, "dim4", 0.5, True, {"applicable": True, "score": 0.5,
                                     "anaphora_type": "single",
                                     "expected_referent": "短袖",
                                     "explanation": "指代部分消解"}),
            (3, "dim5", 1.0, True, {"applicable": True, "score": 1.0,
                                     "expected_theme_source": "user_history",
                                     "explanation": "复述正确"}),
        ]
        for tidx, dcode, score, appl, raw in turn_evals:
            db.add(EvalTurnResult(
                eval_case_result_id=case.id,
                turn_index=tidx,
                dimension_code=dcode,
                score=score,
                applicable=appl,
                judge_raw_response=raw,
            ))
        db.commit()

        yield run.id
    finally:
        # 测试后清理
        db.rollback()
        _cleanup_test_data(db, "v5export-test")
        db.close()


# ---------------------------------------------------------------------------
# 单测
# ---------------------------------------------------------------------------

def test_xlsx_has_5_sheets(synthetic_run):
    db = SessionLocal()
    try:
        data = export_eval_run_xlsx(db, synthetic_run)
    finally:
        db.close()
    wb = load_workbook(io.BytesIO(data))
    assert {"总览", "维度汇总", "Per-Turn 维度明细", "Session 维度明细", "会话明细"}.issubset(
        set(wb.sheetnames)
    ), f"got sheets: {wb.sheetnames}"


def test_per_turn_sheet_has_user_query_and_reasoning(synthetic_run):
    db = SessionLocal()
    try:
        data = export_eval_run_xlsx(db, synthetic_run)
    finally:
        db.close()
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Per-Turn 维度明细"]
    # header
    header = [c.value for c in ws[1]]
    assert "用户query (原始)" in header
    assert "改写query" in header
    assert "bot回复" in header
    assert "改写忠实性 reasoning" in header
    assert "指代消解 reasoning" in header

    # 收集所有行的 user_query 列
    uq_col_idx = header.index("用户query (原始)") + 1
    reasoning_col_idx = header.index("指代消解 reasoning") + 1
    rewrite_col_idx = header.index("改写query") + 1
    bot_resp_col_idx = header.index("bot回复") + 1

    user_queries = []
    rewrites = []
    bot_resps = []
    dim4_reasons = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        user_queries.append(row[uq_col_idx - 1].value)
        rewrites.append(row[rewrite_col_idx - 1].value)
        bot_resps.append(row[bot_resp_col_idx - 1].value)
        dim4_reasons.append(row[reasoning_col_idx - 1].value)

    assert "帮我找优惠" in user_queries
    assert "本人身形偏瘦，帮我搭配夏天短袖" in user_queries
    assert "多推几款" in user_queries
    # turn 3 含 dim4 reasoning
    assert any(r and "指代部分消解" in r for r in dim4_reasons)
    # turn 2 bot_response 应当存在
    assert any(b and "感兴趣" in b for b in bot_resps)


def test_session_sheet_has_dim2_evidence(synthetic_run):
    db = SessionLocal()
    try:
        data = export_eval_run_xlsx(db, synthetic_run)
    finally:
        db.close()
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Session 维度明细"]
    header = [c.value for c in ws[1]]
    assert "跨轮记忆保留 得分" in header
    assert "dim2 bot伪声明 false_inherited" in header
    assert "dim2 漏报约束 missed_constraints" in header
    assert "dim2 precision" in header
    assert "dim2 recall" in header
    assert "纠错响应 适用" in header

    # 第 2 行（第一个 case）应当含 false_inherited 证据
    row_vals = [c.value for c in ws[2]]
    fi_col = header.index("dim2 bot伪声明 false_inherited")
    p_col = header.index("dim2 precision")
    r_col = header.index("dim2 recall")
    assert "身高1.71米" in str(row_vals[fi_col])
    assert row_vals[p_col] == 0.75
    assert row_vals[r_col] == 1.0
